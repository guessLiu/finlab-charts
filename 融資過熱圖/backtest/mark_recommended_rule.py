from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT_NAME = "0050_margin_heat_backtest_report.xlsx"
BEST_ID = "depressed20_immediate"
RULE_SHEET = "0050操作守則"


def find_report() -> Path:
    matches = list(Path("Charts").glob(f"*/backtest/{REPORT_NAME}"))
    if not matches:
        raise FileNotFoundError(REPORT_NAME)
    return matches[0]


def style_header(row):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True, size=12)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def highlight_best_rows(ws, id_header: str, fill_color: str):
    header = [cell.value for cell in ws[1]]
    if id_header not in header:
        return
    id_col = header.index(id_header) + 1
    fill = PatternFill("solid", fgColor=fill_color)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    style_header(ws[1])
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2):
        is_best = row[id_col - 1].value == BEST_ID
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if is_best:
                cell.fill = fill
                cell.font = Font(bold=True, size=14)
        if is_best:
            ws.row_dimensions[row[0].row].height = 34


def autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len("" if cell.value is None else str(cell.value)) for cell in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)


def build_rule_sheet(wb):
    if RULE_SHEET in wb.sheetnames:
        del wb[RULE_SHEET]

    ws = wb.create_sheet(RULE_SHEET, 0)
    blue = PatternFill("solid", fgColor="DDEBF7")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="D9EAD3")
    red = PatternFill("solid", fgColor="F4CCCC")
    dark = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    rows = [
        ("0050 操作守則", ""),
        ("推薦版本", "融資過熱 + 連續 3 週低於週20MA 賣出；融資低迷 <20% 買回"),
        ("初始資金", "100 萬元一次投入 0050"),
        ("持有狀態", "除非觸發出場條件，否則持有 0050"),
        ("出場條件", "同時滿足：1. 融資百分位 >= 80%（過熱）  2. 0050 週收盤連續 3 週低於週20MA"),
        ("買回條件", "空手後，融資百分位第一次 < 20%（低迷）時買回 0050，不額外看價格"),
        ("回測期間", "2009-01-04 ~ 2026-05-10"),
        ("期末資產", "30,084,244"),
        ("CAGR", "21.68%"),
        ("MDD", "-22.04%"),
        ("Sharpe", "1.2884"),
        ("結論", "目前測過的條件中，這版期末資產最高、Sharpe 最高，MDD 仍明顯低於買進持有。"),
    ]

    for r, (label, value) in enumerate(rows, 1):
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
        for c in (1, 2):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True, size=22, color="1F4E78")
    ws["A1"].fill = blue
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    for r in range(2, len(rows) + 1):
        ws.cell(r, 1).font = Font(bold=True, size=16, color="1F4E78")
        ws.cell(r, 2).font = Font(bold=r in {2, 5, 6, 8, 9, 10, 11}, size=16)
        ws.cell(r, 1).fill = yellow if r in {2, 5, 6} else blue
        ws.cell(r, 2).fill = yellow if r in {2, 5, 6} else PatternFill("solid", fgColor="FFFFFF")
        ws.row_dimensions[r].height = 56 if r in {5, 6, 12} else 34

    start = len(rows) + 3
    ws.cell(start, 1, "操作流程")
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=4)
    ws.cell(start, 1).fill = dark
    ws.cell(start, 1).font = Font(color="FFFFFF", bold=True, size=18)
    ws.cell(start, 1).alignment = Alignment(horizontal="center")

    flow = [
        ("狀態", "檢查條件", "動作", "備註"),
        ("持有 0050", "融資 >=80% 且週收盤連 3 週 < 週20MA", "賣出轉現金", "週資料收盤判斷"),
        ("持有 0050", "未同時滿足出場條件", "續抱", "不要因單週過熱就賣"),
        ("空手", "融資百分位第一次 <20%", "全額買回 0050", "不要求站上週20MA"),
        ("空手", "尚未進入 <20%", "續抱現金", "等待低迷區"),
    ]
    for i, row in enumerate(flow, start + 1):
        for j, value in enumerate(row, 1):
            cell = ws.cell(i, j, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if i == start + 1:
                cell.fill = dark
                cell.font = Font(color="FFFFFF", bold=True, size=12)
            else:
                cell.font = Font(size=13, bold=j == 3)
                if j == 3:
                    cell.fill = green if "買回" in value or "續抱" in value else red

    widths = {"A": 22, "B": 64, "C": 24, "D": 34}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def main():
    path = find_report()
    wb = load_workbook(path)

    highlight_best_rows(wb["summary"], "variant_id", "FFF2CC")
    highlight_best_rows(wb["conditions"], "id", "FFF2CC")
    highlight_best_rows(wb["trades"], "variant_id", "D9EAD3")
    build_rule_sheet(wb)

    for ws in wb.worksheets:
        if ws.title != RULE_SHEET:
            autosize(ws)

    wb.save(path)
    print(path)


if __name__ == "__main__":
    main()
